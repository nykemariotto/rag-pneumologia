"""Gera uma planilha Excel do benchmark para validacao medica (revisao/correcao)."""
import os, json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
bm = json.load(open(os.path.join(ROOT, "benchmark", "perguntas.json"), encoding="utf-8"))

wb = Workbook(); ws = wb.active; ws.title = "Benchmark"
headers = ["ID", "Doenca", "Dificuldade", "Pergunta",
           "Resposta-referencia (gabarito)", "Fonte",
           "Correcao / observacoes (preencher)", "OK? (S/N)"]
ws.append(headers)
for c in ws[1]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="2F5496")
    c.alignment = Alignment(wrap_text=True, vertical="center")

cores = {"dpoc": "E8F0FE", "fibrose": "FCE8E6", "pcm": "E6F4EA"}
for q in bm["perguntas"]:
    ws.append([q["id"], q["disease"], q["dificuldade"], q["pergunta"],
               q["resposta_referencia"], q["fonte"], "", ""])
    fill = PatternFill("solid", fgColor=cores.get(q["disease"], "FFFFFF"))
    ws.cell(row=ws.max_row, column=2).fill = fill

widths = {"A": 10, "B": 9, "C": 11, "D": 50, "E": 62, "F": 34, "G": 32, "H": 9}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
for row in ws.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(wrap_text=True, vertical="top")
ws.freeze_panes = "A2"

out = os.path.join(ROOT, "benchmark", "benchmark_para_validar.xlsx")
wb.save(out)
print(f"Salvo: {out}  ({len(bm['perguntas'])} perguntas)")
